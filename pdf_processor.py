#!/usr/bin/env python3
"""
Shared PDF Processing Logic
Core functionality for extracting information from PDFs
"""

import os
import sys
import re
import base64
import io
import difflib

try:
    from openpyxl import load_workbook
except ImportError as e:
    print(f"[WARN] openpyxl not available: {e}")
    load_workbook = None
except Exception as e:
    print(f"[WARN] openpyxl import failed: {e}")
    load_workbook = None

try:
    import fitz  # pymupdf
except ImportError:
    fitz = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import pytesseract
    from pdf2image import convert_from_path
except ImportError:
    pytesseract = None
    convert_from_path = None

# Configure bundled Tesseract/Poppler paths when running from PyInstaller
_BUNDLE_DIR = getattr(sys, "_MEIPASS", None)
_TESSERACT_AVAILABLE = False
_POPPLER_PATH = None  # Store poppler path for pdf2image

if _BUNDLE_DIR:
    _tesseract_exe = os.path.join(_BUNDLE_DIR, "tesseract", "tesseract.exe")
    if os.path.isfile(_tesseract_exe) and pytesseract:
        pytesseract.pytesseract.tesseract_cmd = _tesseract_exe
        os.environ["TESSDATA_PREFIX"] = os.path.join(
            _BUNDLE_DIR, "tesseract", "tessdata"
        )
        _TESSERACT_AVAILABLE = True
        print(f"[DEBUG] Tesseract configured at: {_tesseract_exe}")
    else:
        print(f"[WARNING] Tesseract not found at: {_tesseract_exe}")

    _poppler_dir = os.path.join(_BUNDLE_DIR, "poppler")
    if os.path.isdir(_poppler_dir):
        _POPPLER_PATH = _poppler_dir  # Store for explicit use in convert_from_path
        os.environ["PATH"] = _poppler_dir + os.pathsep + os.environ.get("PATH", "")
        print(f"[DEBUG] Poppler configured at: {_poppler_dir}")
    else:
        print(f"[WARNING] Poppler directory not found at: {_poppler_dir}")
else:
    # Not bundled - check for local installations on Windows
    if sys.platform == "win32":
        import glob

        # Auto-detect Tesseract
        if pytesseract:
            _tesseract_local = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
            if os.path.isfile(_tesseract_local):
                pytesseract.pytesseract.tesseract_cmd = _tesseract_local
            _TESSERACT_AVAILABLE = True

        # Auto-detect Poppler
        _poppler_candidates = glob.glob(r"C:\poppler\poppler-*\Library\bin") + [
            r"C:\poppler\Library\bin",
            r"C:\ProgramData\chocolatey\lib\poppler\tools\Library\bin",
        ]
        for _candidate in _poppler_candidates:
            if os.path.isfile(os.path.join(_candidate, "pdfinfo.exe")):
                _POPPLER_PATH = _candidate
                break
    elif pytesseract:
        _TESSERACT_AVAILABLE = True


class PDFProcessor:
    """Core PDF processing functionality shared between CLI and web app"""

    # ------------------------------------------------------------------ #
    # Pre-compiled regex constants (avoid repeated compilation per call)  #
    # ------------------------------------------------------------------ #
    _RE_SPACES = re.compile(r"\s+")
    _RE_NORM_DATE_STRIP = re.compile(r"[^\d./-]")
    _RE_NORM_DATE_MULTI_DASH = re.compile(r"-{2,}")

    _RE_DATE_LABELED = [
        re.compile(
            r"(?:FECHA\s+DE\s+EVALUACI[OÓ]N|FECHA\s+DE\s+EXAMEN(?:\s+INICIAL)?|"
            r"F\.\s*DE\s+EXAMEN|FECHA\s+EXAMEN|FECHA\s+DE\s+ATENCI[OÓ]N|"
            r"FECHA\s+DE\s+APTITUD)\s*[:\-]?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})",
            re.IGNORECASE,
        ),
        re.compile(
            r"(?:APELLIDOS|NOMBRES|NOMBRE).{0,80}"
            r"FECHA\s*[:\-]\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})",
            re.IGNORECASE,
        ),
        re.compile(
            r"(?<!\w)FECHA\s*[:\-]\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})",
            re.IGNORECASE,
        ),
    ]
    _RE_DATE_ANY = [
        re.compile(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"),
        re.compile(r"\d{1,2}\.\d{1,2}\.\d{2,4}"),
        re.compile(r"\d{4}[/-]\d{1,2}[/-]\d{1,2}"),
    ]

    # Matches "DNI: 12345678" and "DNI O CARNET DE EXTRANJERIA 241953936"
    # Also handles OCR variants: "DNI © CARNET" (© misread as O)
    # Supports 8-digit DNI and 9-12 digit carnet de extranjería
    _RE_DNI_LABELED = re.compile(
        r"\bDNI(?:\s+[O©0o]\s+CARNET\s+DE\s+EXTRANJERIA)?\s*[:\-]?\s*(\d{8,12})\b",
        re.IGNORECASE,
    )
    _RE_DNI_8 = re.compile(r"\b(\d{8})\b")
    _RE_DNI_9PLUS = re.compile(r"\b(\d{9,12})\b")

    _EXAM_RE_STR = (
        r"PRE[- ]?OCUPACIONAL|POST[- ]?OCUPACIONAL|PERI[OÓ]DICO"
        r"|ANUAL|INGRESO|EGRESO|RETIRO"
    )
    _RE_EXAM_CHECKBOX = [
        re.compile(r"(" + _EXAM_RE_STR + r")\s*\|?\s*[xX✓✗☒]\b", re.IGNORECASE),
        re.compile(r"[|]?\s*[xX✓✗☒]\s*\|?\s*(" + _EXAM_RE_STR + r")", re.IGNORECASE),
    ]
    _RE_EXAM_LABELED_LINE = re.compile(
        r"TIPO\s+DE\s+(?:EXAMEN|EVALUACI[OÓ]N)\s*[:\-]\s*(" + _EXAM_RE_STR + r")",
        re.IGNORECASE,
    )
    _RE_EXAM_CONTEXTUAL = re.compile(
        r"EXAMEN\s+M[EÉ]DICO\s+(" + _EXAM_RE_STR + r")",
        re.IGNORECASE,
    )

    _RE_PERSON_LABELED = re.compile(
        r"(?:APELLIDOS?\s+Y\s+NOMBRES?|NOMBRES?\s+Y\s+APELLIDOS?|"
        r"APELLIDOS\s+Y\s+NOMBRE|NOMBRE\s+COMPLETO)[.\s]*[:\-]?\s*"
        r"([A-ZÁÉÍÓÚÑa-záéíóúñ\s]{5,80})",
        re.IGNORECASE,
    )
    _RE_PERSON_3WORDS = re.compile(
        r"\b([A-ZÁÉÍÓÚÑ]{2,25}(?:\s+[A-ZÁÉÍÓÚÑ]{2,25}){2,4})\b"
    )

    _RE_COMPANY_LABELED = re.compile(
        r"(?:EMPRESA|RAZON\s+SOCIAL|RAZÓN\s+SOCIAL|CONTRATISTA|CLIENTE|"
        r"COMPAÑIA|COMPAÑÍA|COMPANIA|COMPANY)(?:\s*[/]\s*\w+)*\s*[:\-|]?\s*"
        r"([A-ZÁÉÍÓÚÑ0-9a-záéíóúñ&\s.]{3,120})",
        re.IGNORECASE,
    )
    _RE_COMPANY_OCR_AMP1 = re.compile(r"(\w)\s*á\s+(\w)")
    _RE_COMPANY_OCR_AMP2 = re.compile(r"&\$")
    _RE_COMPANY_OCR_AMP3 = re.compile(r"\s*&\s*")

    _RE_HUDBAY = [
        re.compile(r"H\s*U\s*D\s*B\s*A\s*Y"),
        re.compile(r"FOR-SS[O0]-\d{3}"),
        re.compile(r"FORMATOS\s+PARA\s+LA\s+VALORACI[OÓ]N\s+DE\s+LA\s+APTITUD"),
        re.compile(r"AUTORIZADO\s+POR\s+HUDBAY"),
    ]

    _RE_FILENAME_ILLEGAL = re.compile(r'[<>:"/\\|?*]')

    @staticmethod
    def _dedupe_keep_order(values):
        seen = set()
        out = []
        for v in values:
            if not v:
                continue
            if v in seen:
                continue
            seen.add(v)
            out.append(v)
        return out

    @classmethod
    def _clean_spaces(cls, s: str) -> str:
        return cls._RE_SPACES.sub(" ", (s or "").strip())

    @classmethod
    def _normalize_date(cls, s: str) -> str:
        """
        Normaliza fechas a un formato consistente usando '-' como separador.
        Ejemplos: 31.12.25 -> 31-12-25 ; 31/12/2025 -> 31-12-2025
        """
        if not s:
            return ""
        cleaned = cls._RE_NORM_DATE_STRIP.sub("", s.strip())
        cleaned = cleaned.replace(".", "-").replace("/", "-")
        cleaned = cls._RE_NORM_DATE_MULTI_DASH.sub("-", cleaned)
        return cleaned.strip("-")

    def extract_text_from_pdf(self, pdf_path, use_ocr=True, verbose=False, pages=None):
        """Extrae texto de PDF (digital o escaneado).

        Args:
            pages: Optional list of 0-indexed page numbers to extract from.
                   If None, extracts from all pages (original behavior).
        """
        text = ""

        # Intenta extraer texto digital primero
        if fitz is not None:
            try:
                doc = fitz.open(pdf_path)
                if pages is not None:
                    for page_num in pages:
                        if page_num < len(doc):
                            text += doc[page_num].get_text()
                    if verbose:
                        print(
                            f"  [INFO] Extrayendo texto digital de paginas: {[p+1 for p in pages]}"
                        )
                else:
                    for page in doc:
                        text += page.get_text()
                doc.close()

                # Si hay suficiente texto, no necesita OCR
                if len(text.strip()) > 50:
                    return text
            except Exception as e:
                if verbose:
                    print(f"  [WARN] Error extrayendo texto digital: {e}")

        # Si no hay texto o es muy poco, usa OCR
        if use_ocr and pytesseract is not None and convert_from_path is not None:
            if verbose:
                print(f"  [OCR] Aplicando OCR (documento escaneado)...")
            try:
                if pages is not None:
                    # Convert only specific pages
                    images = []
                    for page_num in pages:
                        page_images = convert_from_path(
                            pdf_path,
                            first_page=page_num + 1,  # pdf2image uses 1-indexed
                            last_page=page_num + 1,
                            dpi=300,
                            poppler_path=_POPPLER_PATH,
                        )
                        images.extend(page_images)
                    if verbose:
                        print(f"  [INFO] OCR de paginas: {[p+1 for p in pages]}")
                else:
                    images = convert_from_path(
                        pdf_path,
                        first_page=1,
                        last_page=3,
                        dpi=300,
                        poppler_path=_POPPLER_PATH,
                    )  # Solo primeras 3 páginas (fallback)
                if verbose:
                    print(f"  [OK] PDF convertido a {len(images)} imagenes")

                # Then apply OCR to each image (requires Tesseract)
                # Prefer Spanish for ñ/accents; fall back to eng if spa unavailable
                ocr_lang = "spa+eng"
                try:
                    available = pytesseract.get_languages()
                    if "spa" not in available:
                        ocr_lang = "eng"
                except Exception:
                    pass
                # --psm 6: assume uniform block of text (best for form documents)
                ocr_config = "--psm 6"
                for i, image in enumerate(images):
                    ocr_text = pytesseract.image_to_string(
                        image, lang=ocr_lang, config=ocr_config
                    )
                    text += ocr_text
                    if verbose:
                        print(
                            f"  [OK] OCR pagina {i+1}: {len(ocr_text)} caracteres extraidos"
                        )
            except Exception as e:
                if verbose:
                    print(f"  [ERROR] Error en OCR: {type(e).__name__}: {e}")
                    import traceback

                    traceback.print_exc()
                return ""

        return text

    def extract_dates(self, text):
        """Extrae fechas del texto"""
        date_patterns = [
            r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}",  # 31/12/2024 o 31-12-2024
            r"\d{4}[/-]\d{1,2}[/-]\d{1,2}",  # 2024/12/31
            r"\d{1,2}\.\d{1,2}\.\d{2,4}",  # 31.12.2025 o 31.12.25
            r"\d{1,2}\s+(?:de\s+)?"
            r"(?:enero|febrero|marzo|abril|mayo|junio|julio|"
            r"agosto|septiembre|octubre|noviembre|diciembre)"
            r"\s+(?:de\s+)?\d{4}",
        ]

        dates = []
        for pattern in date_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            dates.extend(matches)

        # Prefiere fechas más recientes o las que aparecen primero
        return dates[0] if dates else None

    def extract_numbers(self, text):
        """Extrae números de referencia, factura, DNI, etc."""
        patterns = [
            r"DNI\s*[:\-]?\s*(\d{8})",  # DNI peruano (8 dígitos)
            r"DNI\s*[:\-]?\s*(\d+)",  # DNI general
            r"(?:N°|Nº|No\.|Número|Number|#)\s*[:\-]?\s*(\d+)",
            r"(?:Factura|Invoice|Boleta|Recibo|Receipt)\s*[:\-]?\s*[A-Z]?\d+",
            r"(?:RUC|RFC|NIT|CI)\s*[:\-]?\s*\d+",
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                # Si capturó un grupo, usa ese; si no, usa todo el match
                result = match.group(1) if match.lastindex else match.group(0)
                return result.replace(" ", "_")

        return None

    def extract_entity_names(self, text):
        """Extrae nombres de empresas o personas"""
        lines = text.split("\n")[:30]  # Primeras 30 líneas

        # Patrones para nombres de personas (apellidos y nombres)
        person_patterns = [
            r"(?:APELLIDOS?\s+Y\s+NOMBRES?|NOMBRES?\s+Y\s+APELLIDOS?|NOMBRE|NAME)"
            r"[:\s]+([A-ZÁÉÍÓÚÑ\s]{10,60})",
            # 3 palabras en mayúsculas
            r"([A-ZÁÉÍÓÚÑ]{2,20}\s+[A-ZÁÉÍÓÚÑ]{2,20}\s+[A-ZÁÉÍÓÚÑ]{2,20})",
        ]

        # Busca nombres de personas primero
        for pattern in person_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                name = (
                    match.group(1).strip()
                    if match.lastindex
                    else match.group(0).strip()
                )
                # Limpia y valida
                clean_name = re.sub(r"[^\w\sÁÉÍÓÚÑáéíóúñ]", "", name)
                if 10 < len(clean_name) < 60 and len(clean_name.split()) >= 2:
                    return clean_name.strip()

        # Si no encuentra persona, busca empresas
        for line in lines:
            line = line.strip()
            # Busca líneas con palabras en mayúsculas (empresas)
            if 5 < len(line) < 80 and any(c.isupper() for c in line):
                # Evita líneas que son solo números o fechas
                if re.search(r"[A-ZÁÉÍÓÚÑ]{3,}", line):
                    clean_name = re.sub(r"[^\w\sÁÉÍÓÚÑáéíóúñ&]", "", line)
                    if clean_name and 5 < len(clean_name) < 80:
                        # Limita a palabras razonables
                        words = clean_name.split()
                        if 1 <= len(words) <= 6:
                            return " ".join(words).strip()

        return None

    def detect_document_type(self, text):
        """Detecta el tipo de documento"""
        text_lower = text.lower()

        doc_types = {
            "emoa": [
                "emoa",
                "evaluación médica ocupacional",
                "evaluacion medica ocupacional",
                "examen médico ocupacional",
            ],
            "factura": ["factura", "invoice", "bill"],
            "recibo": ["recibo", "receipt"],
            "contrato": ["contrato", "contract", "acuerdo"],
            "certificado": ["certificado", "certificate", "constancia"],
            "boleta": ["boleta", "ticket"],
            "orden": ["orden de compra", "purchase order"],
            "reporte": ["reporte", "report", "informe"],
            "licencia": ["licencia de conducir", "licencia"],
        }

        for doc_type, keywords in doc_types.items():
            if any(keyword in text_lower for keyword in keywords):
                return doc_type.upper() if doc_type == "emoa" else doc_type.capitalize()

        return None

    def extract_exam_type(self, text):
        """Extrae tipo de examen médico (PERIODICO, INGRESO, etc.)"""
        text_upper = text.upper()
        exam_types = [
            "PERIODICO",
            "INGRESO",
            "EGRESO",
            "RETIRO",
            "PREOCUPACIONAL",
            "POSTOCUPACIONAL",
        ]

        for exam_type in exam_types:
            if exam_type in text_upper:
                return exam_type

        return None

    def extract_date_candidates(self, text: str):
        """Devuelve lista de fechas (normalizadas) encontradas en el texto.
        Prioritizes dates with labels like 'FECHA DE EXAMEN', 'FECHA DE EVALUACION'.
        """
        if not text:
            return []

        # 1. Try labeled dates first (highest priority)
        labeled_dates = []
        for pat in self._RE_DATE_LABELED:
            for m in pat.finditer(text):
                labeled_dates.append(self._normalize_date(m.group(1)))

        # 2. All dates as fallback
        all_dates = []
        for pat in self._RE_DATE_ANY:
            all_dates.extend(pat.findall(text))

        # Labeled dates first, then all others
        combined = labeled_dates + [self._normalize_date(d) for d in all_dates]
        return self._dedupe_keep_order(combined)

    def extract_dni_candidates(self, text: str):
        """Devuelve lista de posibles DNI/carnet encontrados en el texto.

        Handles:
        - 8-digit Peruvian DNI:  "DNI: 76248882"
        - 9-12 digit carnet:     "DNI O CARNET DE EXTRANJERIA 241953936"

        When OCR produces a 9-digit number in a labeled DNI field, also adds
        the 8-digit variant (first digit stripped) as a candidate, since OCR
        sometimes prepends a stray character to the actual 8-digit DNI.
        """
        if not text:
            return []
        found = []
        # Labeled first (highest confidence — includes carnet de extranjería)
        for match in self._RE_DNI_LABELED.findall(text):
            found.append(match)
            # If 9 digits, also offer the 8-digit variant (OCR artifact guard)
            if len(match) == 9:
                found.append(match[1:])
        # Fallback: bare 8-digit blocks
        found.extend(self._RE_DNI_8.findall(text))
        # Fallback: bare 9-12 digit blocks only if nothing found yet
        if not found:
            found.extend(self._RE_DNI_9PLUS.findall(text))
        return self._dedupe_keep_order(found)

    # Map checkbox/form labels to canonical exam types
    _EXAM_LABEL_MAP = {
        "PREOCUPACIONAL": "PREOCUPACIONAL",
        "PRE-OCUPACIONAL": "PREOCUPACIONAL",
        "PRE OCUPACIONAL": "PREOCUPACIONAL",
        "POSTOCUPACIONAL": "POSTOCUPACIONAL",
        "POST-OCUPACIONAL": "POSTOCUPACIONAL",
        "POST OCUPACIONAL": "POSTOCUPACIONAL",
        "PERIODICO": "PERIODICO",
        "PERIÓDICO": "PERIODICO",
        "ANUAL": "PERIODICO",
        "INGRESO": "INGRESO",
        "EGRESO": "EGRESO",
        "RETIRO": "RETIRO",
    }

    def extract_exam_type_candidates(self, text: str):
        """Devuelve lista de tipos de examen encontrados en el texto.

        Prioritizes:
        1. Labeled fields like 'TIPO DE EXAMEN: PREOCUPACIONAL'
        2. Checkbox-style forms where 'x' marks the selected type
        3. Contextual phrases like 'EXAMEN MÉDICO PERIODICO'
        4. Fallback: any exam type keyword found in text
        """
        if not text:
            return []

        prioritized = []

        # 1. Checkbox-style: "Anual x" or "x Periodico" (x marks the checked option)
        #    Highest priority because it indicates what's actually selected.
        for pat in self._RE_EXAM_CHECKBOX:
            for m in pat.finditer(text):
                canonical = self._EXAM_LABEL_MAP.get(m.group(1).upper().strip())
                if canonical:
                    prioritized.append(canonical)

        # 2. Labeled on same line: "TIPO DE EXAMEN: PREOCUPACIONAL"
        for line in text.split("\n"):
            for match in self._RE_EXAM_LABELED_LINE.findall(line):
                canonical = self._EXAM_LABEL_MAP.get(match.upper().strip())
                if canonical:
                    prioritized.append(canonical)

        # 3. Contextual: "EXAMEN MÉDICO PERIODICO"
        for match in self._RE_EXAM_CONTEXTUAL.findall(text):
            canonical = self._EXAM_LABEL_MAP.get(match.upper().strip())
            if canonical:
                prioritized.append(canonical)

        if prioritized:
            return self._dedupe_keep_order(prioritized)

        # 4. Fallback: any exam type keyword in text
        upper = text.upper()
        normalized = (
            upper.replace("PRE-OCUPACIONAL", "PREOCUPACIONAL")
            .replace("POST-OCUPACIONAL", "POSTOCUPACIONAL")
            .replace("PRE OCUPACIONAL", "PREOCUPACIONAL")
            .replace("POST OCUPACIONAL", "POSTOCUPACIONAL")
        )
        exam_types = [
            "PREOCUPACIONAL",
            "POSTOCUPACIONAL",
            "PERIODICO",
            "INGRESO",
            "EGRESO",
            "RETIRO",
        ]
        found = [t for t in exam_types if t in normalized]
        return self._dedupe_keep_order(found)

    # Words that should not appear in a person's name (noise from OCR)
    _NAME_NOISE_WORDS = {
        "AREA",
        "DNI",
        "CARGO",
        "PUESTO",
        "FECHA",
        "EMPRESA",
        "RUC",
        "TELEFONO",
        "CELULAR",
        "CORREO",
        "EMAIL",
        "DIRECCION",
        "DISTRITO",
        "PROVINCIA",
        "DEPARTAMENTO",
        "PERU",
        "LIMA",
        "CARNET",
        "EXTRANJERIA",
        "DOCUMENTO",
        "IDENTIDAD",
        "TRABAJADOR",
        "PACIENTE",
        "EVALUADO",
        "EXAMINADO",
        "CONTRATA",
        "CONTRATISTA",
        "SAC",
        "SRL",
        "EIRL",
        "OCUPACIONAL",
        "MEDICO",
        "EXAMEN",
        "RESULTADO",
        "INGRESO",
        "EGRESO",
        "PERIODICO",
        "PREOCUPACIONAL",
        "POSTOCUPACIONAL",
        "RETIRO",
        "TIPO",
        "EVALUACION",
        "FORMATOS",
        "PARA",
        "CONSENTIMIENTO",
        "INFORMADO",
        "NUMERO",
        "PASAPORTE",
        "SERVICIOS",
        "LOGISTICA",
        "INFORME",
        "LLENADO",
    }

    def _clean_person_name(self, raw: str) -> str:
        """Clean a raw person name: remove noise words, limit to 5 words."""
        words = self._clean_spaces(raw).split()
        clean = []
        for w in words:
            w_upper = w.upper().strip(".,;:()")
            if w_upper in self._NAME_NOISE_WORDS:
                break  # stop at first noise word
            if len(w_upper) < 2:
                continue
            if w_upper.isdigit():
                break
            clean.append(w_upper)
            if len(clean) >= 5:
                break
        return " ".join(clean)

    def extract_person_name_candidates(self, text: str):
        """Devuelve lista de posibles nombres de persona encontrados en el texto."""
        if not text:
            return []
        candidates = []

        for m in self._RE_PERSON_LABELED.finditer(text):
            raw = m.group(1).split("\n", 1)[0].strip()
            cleaned = self._clean_person_name(raw)
            if len(cleaned.split()) >= 2:
                candidates.append(cleaned)

        # Heurística: 3 palabras mayúsculas consecutivas (típico en reportes)
        for m in self._RE_PERSON_3WORDS.finditer(text):
            raw = self._clean_spaces(m.group(1))
            # evita empresas u otros
            if any(
                k in raw.upper()
                for k in [
                    "CONSORCIO",
                    "EMPRESA",
                    "CONTRATISTA",
                    "SAC",
                    "S.A",
                    "S.A.C",
                    "S.R.L",
                    "CENTRO",
                    "MEDICO",
                ]
            ):
                continue
            cleaned = self._clean_person_name(raw)
            if len(cleaned.split()) >= 2:
                candidates.append(cleaned)

        return self._dedupe_keep_order(candidates)

    def _clean_company_name(self, raw: str) -> str:
        """Clean company name: keep legal suffixes, limit to 5 words, filter OCR noise."""
        raw = self._clean_spaces(raw)
        # Fix common OCR misreads of '&': "Má" -> "M&", "&$" -> "&", collapse spaces around &
        raw = self._RE_COMPANY_OCR_AMP1.sub(r"\1&\2", raw)  # "Má S" -> "M&S"
        raw = self._RE_COMPANY_OCR_AMP2.sub("&", raw)  # "&$" -> "&"
        raw = self._RE_COMPANY_OCR_AMP3.sub("&", raw)  # collapse spaces around &
        raw = self._clean_spaces(raw)
        # Reject if mostly lowercase (OCR garbage like "contiene informacién...")
        words = raw.split()
        upper_count = sum(1 for w in words if w[0].isupper() if w)
        if len(words) > 3 and upper_count < len(words) * 0.5:
            return ""
        # Limit to 5 words
        words = words[:5]
        return " ".join(words)

    def extract_company_candidates(self, text: str):
        """Devuelve lista de posibles empresas encontradas en el texto."""
        if not text:
            return []
        candidates = []

        for m in self._RE_COMPANY_LABELED.finditer(text):
            raw = m.group(1).split("\n", 1)[0].strip()
            cleaned = self._clean_company_name(raw)
            if len(cleaned) >= 3:
                candidates.append(cleaned)

        # Heurística: líneas con CONSORCIO
        for line in text.split("\n")[:80]:
            ln = self._clean_spaces(line)
            if not ln:
                continue
            if "CONSORCIO" in ln.upper():
                cleaned = self._clean_company_name(ln)
                if cleaned:
                    candidates.append(cleaned)

        return self._dedupe_keep_order(candidates)

    # Map full exam type names to abbreviations
    _EXAM_TYPE_ABBR = {
        "PREOCUPACIONAL": "EMPO",
        "PERIODICO": "EMOA",
        "POSTOCUPACIONAL": "EMOR",
        "INGRESO": "INGRESO",
        "EGRESO": "EGRESO",
        "RETIRO": "RETIRO",
    }

    @staticmethod
    def _date_to_short(date_str: str) -> str:
        """Convert normalized date (DD-MM-YYYY) to short format (DD.MM.YY)."""
        if not date_str:
            return ""
        parts = date_str.replace("/", "-").replace(".", "-").split("-")
        if len(parts) != 3:
            return date_str.replace("-", ".")
        dd, mm, yy = parts[0], parts[1], parts[2]
        # Shorten year to 2 digits if 4 digits
        if len(yy) == 4:
            yy = yy[2:]
        return f"{dd}.{mm}.{yy}"

    def build_filename(
        self,
        dni="",
        nombre="",
        empresa="",
        tipo_examen="",
        fecha="",
        include=None,
        fmt="hudbay",
    ):
        """
        Construye el nombre final.

        fmt="hudbay":   FECHA TIPO DNI NOMBRE-EMPRESA.pdf
                        31.01.26 EMPO 76248882 HUAMAN POCCO JESUS YOVANI-G4S PERU SAC.pdf

        fmt="standard": DNI-NOMBRE-EMPRESA-TIPO-CMESPINAR-FECHA.pdf
                        45205399-INFANTE CHUQUIRUNA JULIO CESAR-KOMATSU MITSUI
                        MAQUINARIAS PERU S.A.-EMPO-CMESPINAR-02.02.26.pdf
        """
        include = include or {}

        def use(field, value):
            if field in include:
                return bool(include[field])
            return True

        dni = self._clean_spaces(dni)
        nombre = self._clean_spaces(nombre).upper()
        empresa = self._clean_spaces(empresa).upper()
        tipo_examen = self._clean_spaces(tipo_examen).upper()
        fecha = self._normalize_date(fecha)

        # Abbreviate exam type
        tipo_abbr = self._EXAM_TYPE_ABBR.get(tipo_examen, tipo_examen)

        if fmt == "standard":
            # Standard: DNI-NOMBRE-EMPRESA-TIPO-CMESPINAR-FECHA.pdf
            parts = []
            if use("dni", dni) and dni:
                parts.append(dni)
            if use("nombre", nombre) and nombre:
                parts.append(nombre)
            if use("empresa", empresa) and empresa:
                parts.append(empresa)
            if use("tipo_examen", tipo_abbr) and tipo_abbr:
                parts.append(tipo_abbr)
            parts.append("CMESPINAR")
            if use("fecha", fecha) and fecha:
                parts.append(self._date_to_short(fecha))
            new_name = "-".join([p for p in parts if p])
            new_name = self._RE_FILENAME_ILLEGAL.sub("", new_name)
            return new_name + ".pdf" if new_name else ""

        # Hudbay (default): FECHA TIPO DNI NOMBRE-EMPRESA.pdf
        parts = []

        if use("fecha", fecha) and fecha:
            parts.append(self._date_to_short(fecha))
        if use("tipo_examen", tipo_abbr) and tipo_abbr:
            parts.append(tipo_abbr)
        if use("dni", dni) and dni:
            parts.append(dni)

        # Nombre and empresa joined with hyphen
        name_empresa = ""
        if use("nombre", nombre) and nombre:
            name_empresa = nombre
        if use("empresa", empresa) and empresa:
            if name_empresa:
                name_empresa += "-" + empresa
            else:
                name_empresa = empresa
        if name_empresa:
            parts.append(name_empresa)

        new_name = " ".join([p for p in parts if p])
        new_name = self._RE_FILENAME_ILLEGAL.sub("", new_name)
        return new_name + ".pdf" if new_name else ""

    @classmethod
    def detect_format_from_content(cls, text):
        """Detect if PDF content is a Hudbay document based on OCR text patterns.

        Looks for Hudbay logo OCR variants, document IDs, and explicit mentions.
        Returns 'hudbay' if detected, None otherwise.
        """
        if not text:
            return None
        upper = text.upper()
        for pat in cls._RE_HUDBAY:
            if pat.search(upper):
                return "hudbay"
        return None

    @staticmethod
    def detect_format(filename):
        """Detect if a filename matches Hudbay or Standard format.

        Hudbay:   starts with date (DD.MM.YY), spaces between fields
        Standard: starts with DNI (8 digits), hyphens between fields, contains CMESPINAR

        Returns 'hudbay', 'standard', or None if unrecognized.
        """
        if not filename:
            return None
        name = filename.rsplit(".pdf", 1)[0].rsplit(".PDF", 1)[0].strip()
        if not name:
            return None

        # Standard: starts with 8-digit DNI, uses hyphens, contains CMESPINAR
        if re.match(r"^\d{8}-", name) and "CMESPINAR" in name.upper():
            return "standard"

        # Hudbay: starts with date DD.MM.YY, uses spaces
        if re.match(r"^\d{1,2}\.\d{1,2}\.\d{2,4}\s", name):
            return "hudbay"

        return None

    @staticmethod
    def get_excel_info(excel_path):
        """Return all sheet names and their column headers.

        Returns {sheet_name: [col1, col2, ...]} for every sheet in the workbook.
        """
        if load_workbook is None:
            raise ImportError("openpyxl is required to read Excel files.")
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        info = {}
        for name in wb.sheetnames:
            ws = wb[name]
            header_row = next(
                ws.iter_rows(min_row=1, max_row=1, values_only=True), None
            )
            info[name] = [str(h).strip() if h else "" for h in (header_row or [])]
        wb.close()
        return info

    @staticmethod
    def load_excel_reference(
        excel_path,
        sheet_name=None,
        dni_col=None,
        hudbay_col=None,
        standard_col=None,
    ):
        """Load an Excel file and build a DNI-keyed lookup dict.

        Args:
            excel_path: path to the .xlsx file
            sheet_name: sheet to read; None → try 'CARGAS EN MEDIWEB', then active sheet
            dni_col:      exact column header for the DNI/document field (case-insensitive)
            hudbay_col:   exact column header for the Hudbay reference filename
            standard_col: exact column header for the standard reference filename

        Returns dict {dni_string: {"paciente": str|None,
        "hudbay_name": str|None, "standard_name": str|None}}.
        """
        if load_workbook is None:
            raise ImportError(
                "openpyxl is required to read Excel files. Install with: pip install openpyxl"
            )

        wb = load_workbook(excel_path, read_only=True, data_only=True)

        # Sheet selection
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif "CARGAS EN MEDIWEB" in wb.sheetnames:
            ws = wb["CARGAS EN MEDIWEB"]
        else:
            ws = wb.active

        # Build header index (raw for matching, lower for auto-detect)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers_raw = [str(h).strip() if h else "" for h in header_row]
        headers = [h.lower() for h in headers_raw]

        def _find_col(explicit, *auto_patterns):
            """Return column index by explicit name first, then pattern auto-detect."""
            if explicit:
                for i, h in enumerate(headers_raw):
                    if h.lower() == explicit.lower():
                        return i
            for pat in auto_patterns:
                for i, h in enumerate(headers):
                    if pat in h:
                        return i
            return None

        doc_idx = _find_col(dni_col, "documento", "dni")
        pac_idx = _find_col(None, "paciente", "nombre", "name")
        hudbay_idx = _find_col(
            hudbay_col, "emo hudbay", "hudbay", "nombre excel", "nombre_excel"
        )
        standard_idx = _find_col(
            standard_col,
            "emo bambas",
            "bambas",
            "nombre excel",
            "nombre_excel",
        )

        if doc_idx is None:
            wb.close()
            raise ValueError(
                f"No se encontró columna de DNI/documento en la hoja '{ws.title}'. "
                f"Columnas encontradas: {headers_raw}"
            )

        def _cell(row, idx):
            if idx is not None and idx < len(row) and row[idx] is not None:
                v = row[idx]
                # Excel stores integers as floats (e.g. 76248882.0 → "76248882")
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                return str(v).strip() or None
            return None

        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            doc_val = _cell(row, doc_idx)
            if not doc_val:
                continue
            lookup[doc_val] = {
                "paciente": _cell(row, pac_idx),
                "hudbay_name": _cell(row, hudbay_idx),
                "standard_name": _cell(row, standard_idx),
            }

        wb.close()
        return lookup

    @staticmethod
    def match_name_from_excel(extracted_name, dni, excel_lookup):
        """Try to match a name from the Excel lookup by DNI.

        If the DNI is found in excel_lookup and the similarity between
        extracted_name and the Excel paciente is >= 0.80, returns the
        Excel paciente value. Otherwise returns the original extracted_name.
        """
        if not excel_lookup or not dni:
            return extracted_name

        excel_name = excel_lookup.get(dni)
        if not excel_name:
            return extracted_name

        # Compare similarity
        if not extracted_name:
            return excel_name

        ratio = difflib.SequenceMatcher(
            None,
            extracted_name.upper(),
            excel_name.upper(),
        ).ratio()

        if ratio >= 0.80:
            return excel_name

        return extracted_name

    @staticmethod
    def _find_excel_by_name(extracted_name, excel_lookup, threshold=0.72):
        """Fallback: search all Excel entries by paciente name similarity.

        Returns (entry, ratio) for the best match above threshold, or (None, 0).
        Used when DNI is missing or not found in the Excel file.
        """
        if not extracted_name or not excel_lookup:
            return None, 0
        name_upper = extracted_name.upper()
        best_entry = None
        best_ratio = 0.0
        for entry in excel_lookup.values():
            pac = (entry.get("paciente") or "").upper()
            if not pac:
                continue
            ratio = difflib.SequenceMatcher(None, name_upper, pac).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_entry = entry
        if best_ratio >= threshold:
            return best_entry, best_ratio
        return None, 0.0

    def analyze(
        self,
        pdf_path,
        original_filename=None,
        verbose=False,
        excel_lookup=None,
        forced_format=None,
    ):
        """
        Analiza un PDF y devuelve candidatos + valores por defecto para armar el nombre en UI.
        forced_format: If set ('hudbay' or 'standard'), skip auto-detection and use this format.
        """
        if verbose:
            ocr_status = (
                "disponible" if (pytesseract and convert_from_path) else "NO disponible"
            )
            print(f"  [INFO] OCR {ocr_status}")

        # Determine page-specific extraction based on format
        if forced_format:
            detected_fmt = forced_format
            if verbose:
                print(f"  [INFO] Formato forzado por usuario: {forced_format}")
        else:
            detected_fmt = self.detect_format(original_filename)

        text = None
        if detected_fmt == "hudbay":
            text = self.extract_text_from_pdf(pdf_path, verbose=verbose, pages=[0])
            if verbose:
                print(f"  [INFO] Formato Hudbay -> extrayendo solo pagina 1")
        elif detected_fmt == "standard":
            text = self.extract_text_from_pdf(pdf_path, verbose=verbose, pages=[1])
            if verbose:
                print(f"  [INFO] Formato Estandar -> extrayendo solo pagina 2")
        else:
            # Unknown filename: read page 1 first to detect format from content
            page1_text = self.extract_text_from_pdf(
                pdf_path, verbose=verbose, pages=[0]
            )
            content_fmt = self.detect_format_from_content(page1_text)
            if content_fmt == "hudbay":
                detected_fmt = "hudbay"
                text = page1_text
                if verbose:
                    print(
                        f"  [INFO] Formato Hudbay detectado por contenido -> usando pagina 1"
                    )
            else:
                detected_fmt = "standard"
                text = self.extract_text_from_pdf(pdf_path, verbose=verbose, pages=[1])
                if verbose:
                    print(
                        f"  [INFO] Formato Estandar (por defecto) -> extrayendo solo pagina 2"
                    )
        filename_data = (
            self.extract_from_filename(original_filename) if original_filename else {}
        )

        # candidatos desde texto
        dni_c = self.extract_dni_candidates(text)
        nombre_c = self.extract_person_name_candidates(text)
        empresa_c = self.extract_company_candidates(text)
        tipo_c = self.extract_exam_type_candidates(text)
        fecha_c = self.extract_date_candidates(text)

        # merge con filename_data (fallback)
        if filename_data.get("dni"):
            dni_c.append(filename_data["dni"])
        if filename_data.get("person_name"):
            nombre_c.append(filename_data["person_name"])
        if filename_data.get("company"):
            empresa_c.append(filename_data["company"])
        if filename_data.get("exam_type"):
            tipo_c.append(filename_data["exam_type"])
        if filename_data.get("date"):
            fecha_c.append(self._normalize_date(filename_data["date"]))

        candidates = {
            "dni": self._dedupe_keep_order(dni_c),
            "nombre": self._dedupe_keep_order(
                [self._clean_spaces(x) for x in nombre_c]
            ),
            "empresa": self._dedupe_keep_order(
                [self._clean_spaces(x) for x in empresa_c]
            ),
            "tipo_examen": self._dedupe_keep_order(
                [self._clean_spaces(x).upper() for x in tipo_c]
            ),
            "fecha": self._dedupe_keep_order(
                [self._normalize_date(x) for x in fecha_c]
            ),
        }

        # Excel cross-reference: lookup returns {dni: {"paciente", "hudbay_name", "standard_name"}}
        nombre_excel = None
        excel_dni_found = None  # None = Excel not loaded; True/False when loaded
        if excel_lookup:
            excel_entry = None
            matched_by_name = False

            # Primary match: by DNI
            if candidates["dni"]:
                first_dni = candidates["dni"][0]
                excel_entry = excel_lookup.get(first_dni)
                excel_dni_found = excel_entry is not None
            else:
                excel_dni_found = False

            # Fallback: by name similarity when DNI produced no entry
            if not excel_entry and candidates["nombre"]:
                fallback_entry, fallback_ratio = self._find_excel_by_name(
                    candidates["nombre"][0], excel_lookup
                )
                if fallback_entry:
                    excel_entry = fallback_entry
                    excel_dni_found = True
                    matched_by_name = True
                    if verbose:
                        print(
                            f"  [EXCEL] Coincidencia por nombre"
                            f" ({round(fallback_ratio * 100)}%): "
                            f"'{candidates['nombre'][0]}'"
                        )

            if excel_entry:
                pac_name = excel_entry.get("paciente") or ""
                if pac_name:
                    pac_clean = self._clean_spaces(pac_name)
                    # Always prepend the Excel paciente name as the top candidate
                    candidates["nombre"] = self._dedupe_keep_order(
                        [pac_clean] + candidates["nombre"]
                    )
                    if verbose:
                        src = (
                            "nombre"
                            if matched_by_name
                            else f"DNI: {candidates['dni'][0]}"
                        )
                        print(f"  [EXCEL] Nombre Excel: '{pac_clean}' ({src})")
                # Pick format-specific reference name
                if detected_fmt == "hudbay":
                    nombre_excel = excel_entry.get("hudbay_name") or excel_entry.get(
                        "standard_name"
                    )
                else:
                    nombre_excel = excel_entry.get("standard_name") or excel_entry.get(
                        "hudbay_name"
                    )
                if nombre_excel and verbose:
                    print(f"  [EXCEL] Referencia ({detected_fmt}): '{nombre_excel}'")

        if verbose:
            print(f"\n[DEBUG] Candidatos encontrados:")
            for key, values in candidates.items():
                print(f"  - {key}: {values}")

        defaults = {k: (v[0] if v else "") for k, v in candidates.items()}

        if verbose:
            print(f"\n[DEBUG] Valores por defecto:")
            for key, value in defaults.items():
                print(f"  - {key}: '{value}'")

        # nombre sugerido (default)
        suggested = ""
        success = True
        notes = []
        if not defaults["dni"]:
            success = False
            notes.append("No se detectó DNI (requerido).")
        if not defaults["fecha"]:
            notes.append("No se detectó fecha de evaluación.")

        if defaults["dni"]:
            suggested = self.build_filename(
                dni=defaults["dni"],
                nombre=defaults["nombre"],
                empresa=defaults["empresa"],
                tipo_examen=defaults["tipo_examen"],
                fecha=defaults["fecha"],
                fmt=detected_fmt or "standard",
            )
        else:
            suggested = ""

        # Calculate match percentage against "nombre excel" column
        match_percentage = None
        if nombre_excel and suggested:
            ratio = difflib.SequenceMatcher(
                None,
                suggested.upper().replace(".PDF", ""),
                nombre_excel.upper().replace(".PDF", ""),
            ).ratio()
            match_percentage = round(ratio * 100, 1)

        # detected_fmt is already set above (from filename or content-based detection)
        return {
            "success": success and bool(suggested),
            "suggested_name": suggested or None,
            "candidates": candidates,
            "defaults": defaults,
            "text_chars": len(text or ""),
            "used_filename_fallback": (
                len((text or "").strip()) < 10 and bool(filename_data)
            ),
            "notes": notes,
            "detected_format": detected_fmt,
            "nombre_excel": nombre_excel,
            "match_percentage": match_percentage,
            "excel_dni_found": excel_dni_found,
        }

    # Reverse map: abbreviation -> full exam type name
    _ABBR_TO_EXAM_TYPE = {v: k for k, v in _EXAM_TYPE_ABBR.items()}

    def _parse_exam_type_token(self, token):
        """Resolve an exam type token (full name or abbreviation) to the full name."""
        upper = token.upper().strip()
        if upper in self._EXAM_TYPE_ABBR:
            return upper  # already full name
        if upper in self._ABBR_TO_EXAM_TYPE:
            return self._ABBR_TO_EXAM_TYPE[upper]  # abbreviation -> full
        return None

    def extract_from_filename(self, filename):
        """Extrae información del nombre del archivo como fallback.

        Handles two known formats:
        - Hudbay:   'DD.MM.YY TIPO DNI NOMBRE-EMPRESA.pdf'
        - Standard: 'DNI-NOMBRE-EMPRESA-TIPO-CMESPINAR-DD.MM.YY.pdf'
        - Unknown:  generic heuristic extraction
        """
        parts = {}
        name_without_ext = filename.rsplit(".pdf", 1)[0].rsplit(".PDF", 1)[0].strip()
        if not name_without_ext:
            return parts

        fmt = self.detect_format(filename)

        if fmt == "hudbay":
            return self._parse_hudbay_filename(name_without_ext)
        elif fmt == "standard":
            return self._parse_standard_filename(name_without_ext)
        else:
            return self._parse_generic_filename(name_without_ext)

    def _parse_hudbay_filename(self, name):
        """Parse Hudbay format: 'DD.MM.YY TIPO DNI NOMBRE-EMPRESA'
        Example: '31.01.26 EMPO 76248882 HUAMAN POCCO JESUS YOVANI-G4S PERU SAC'
        """
        parts = {}

        # Split name from company at the hyphen (last major separator)
        # But the hyphen joins NOMBRE-EMPRESA, so split on first hyphen
        main_part = name
        company_part = ""
        if "-" in name:
            # Find the hyphen that separates person name from company
            main_part, company_part = name.split("-", 1)

        tokens = main_part.split()
        # Expected: [DATE, TIPO, DNI, NAME_WORD1, NAME_WORD2, ...]

        idx = 0
        # 1. Date (DD.MM.YY)
        if idx < len(tokens) and re.match(r"\d{1,2}\.\d{1,2}\.\d{2,4}$", tokens[idx]):
            parts["date"] = tokens[idx]
            idx += 1

        # 2. Exam type (abbreviation or full)
        if idx < len(tokens):
            exam = self._parse_exam_type_token(tokens[idx])
            if exam:
                parts["exam_type"] = exam
                idx += 1

        # 3. DNI (8 digits)
        if idx < len(tokens) and re.match(r"\d{8}$", tokens[idx]):
            parts["dni"] = tokens[idx]
            idx += 1

        # 4. Remaining tokens = person name
        if idx < len(tokens):
            person_name = " ".join(tokens[idx:]).strip()
            if len(person_name.split()) >= 2:
                parts["person_name"] = person_name

        # 5. Company (after the hyphen)
        if company_part.strip():
            parts["company"] = company_part.strip()

        return parts

    def _parse_standard_filename(self, name):
        """Parse Standard format: 'DNI-NOMBRE-EMPRESA-TIPO-CMESPINAR-DD.MM.YY'
        Example: '45205399-INFANTE CHUQUIRUNA JULIO CESAR-KOMATSU MITSUI
        MAQUINARIAS PERU S.A.-EMPO-CMESPINAR-02.02.26'
        """
        parts = {}

        # Split by hyphens, but some fields contain spaces (NOMBRE, EMPRESA)
        # Strategy: walk segments from both ends where we know the structure
        segments = name.split("-")
        if len(segments) < 3:
            return self._parse_generic_filename(name)

        idx_start = 0
        idx_end = len(segments) - 1

        # 1. First segment: DNI (8 digits)
        if re.match(r"\d{8}$", segments[0].strip()):
            parts["dni"] = segments[0].strip()
            idx_start = 1

        # From the end: date, CMESPINAR, exam type
        # Last segment: date (DD.MM.YY)
        if re.match(r"\d{1,2}\.\d{1,2}\.\d{2,4}$", segments[idx_end].strip()):
            parts["date"] = segments[idx_end].strip()
            idx_end -= 1

        # CMESPINAR
        if idx_end >= idx_start and segments[idx_end].strip().upper() == "CMESPINAR":
            idx_end -= 1

        # Exam type (abbreviation or full)
        if idx_end >= idx_start:
            exam = self._parse_exam_type_token(segments[idx_end].strip())
            if exam:
                parts["exam_type"] = exam
                idx_end -= 1

        # Middle segments: NOMBRE and EMPRESA
        # We need to figure out where NOMBRE ends and EMPRESA begins
        middle = segments[idx_start : idx_end + 1]
        if len(middle) >= 2:
            # First middle segment is person name, rest is company
            parts["person_name"] = middle[0].strip()
            parts["company"] = "-".join(middle[1:]).strip()
        elif len(middle) == 1:
            # Only one segment - could be name or company
            val = middle[0].strip()
            if len(val.split()) >= 2 and val[0].isalpha():
                parts["person_name"] = val
            else:
                parts["company"] = val

        return parts

    def _parse_generic_filename(self, name_without_ext):
        """Fallback: generic heuristic extraction for unknown formats."""
        parts = {}

        # Extrae fecha (DD.MM.YY o DD-MM-YY)
        date_match = re.search(r"(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})", name_without_ext)
        if date_match:
            parts["date"] = date_match.group(1)

        # Extrae DNI (8 dígitos seguidos)
        dni_match = re.search(r"(\d{8})", name_without_ext)
        if dni_match:
            parts["dni"] = dni_match.group(1)

        # Busca tipo de examen (full names and abbreviations)
        all_exam_tokens = list(self._EXAM_TYPE_ABBR.keys()) + list(
            self._ABBR_TO_EXAM_TYPE.keys()
        )
        upper_name = name_without_ext.upper()
        for token in all_exam_tokens:
            if token in upper_name:
                exam = self._parse_exam_type_token(token)
                if exam:
                    parts["exam_type"] = exam
                    break

        # Extrae nombre de persona (después de DNI, antes de guion)
        name_patterns = [
            r"\d{8}\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{8,50}?)(?:\s*-)",
            r"(?:DNI|dni|ID|id)[:\s]*\d+\s+([A-ZÁÉÍÓÚÑ\s]{10,50}?)(?:-|CONSORCIO|EMPRESA|COMPANY)",
        ]
        for pattern in name_patterns:
            name_match = re.search(pattern, name_without_ext, re.IGNORECASE)
            if name_match:
                person_name = name_match.group(1).strip()
                if (
                    len(person_name.split()) >= 2
                    and not person_name.replace(" ", "").isdigit()
                ):
                    parts["person_name"] = person_name
                    break

        # Extrae empresa (después del guion)
        if "-" in name_without_ext:
            company_part = name_without_ext.split("-", 1)[1]
            company_clean = re.sub(r"[&]", "Y", company_part)
            company_words = company_clean.split()
            if len(company_words) > 4:
                important_words = [
                    w
                    for w in company_words[:6]
                    if w.upper() not in ["MECANICA", "REVESTIMIENTO", "Y"]
                ]
                company_words = (
                    important_words[:4] if important_words else company_words[:3]
                )
            else:
                company_words = company_words[:4]
            if company_words:
                parts["company"] = " ".join(company_words)

        return parts

    # Field highlight colors
    _FIELD_COLORS = {
        "dni": "#ff6b6b",
        "nombre": "#4ecdc4",
        "empresa": "#45b7d1",
        "tipo_examen": "#f7b731",
        "fecha": "#5f27cd",
    }

    MAX_PREVIEW_PAGES = 1  # Only load the single page used for extraction

    def generate_preview_single_page(
        self, pdf_path, defaults, page_num=0, max_width=900
    ):
        """Render a single PDF page as JPEG and find bounding boxes for field values.

        Returns dict with 'page' (image, width, height, highlights, page number)
        plus 'total_pages' count for pagination UI.
        """
        if not defaults:
            defaults = {}

        result = self._preview_digital_single(pdf_path, defaults, page_num, max_width)
        if result is None:
            result = self._preview_ocr_single(pdf_path, defaults, page_num, max_width)
        return result

    def generate_preview_with_highlights(self, pdf_path, defaults, max_width=500):
        """Render PDF pages as images and find bounding boxes for field values.

        Returns dict with 'pages' list plus 'total_pages' count.
        Limited to MAX_PREVIEW_PAGES for performance.
        """
        if not defaults:
            defaults = {}

        # Try digital first, fall back to OCR
        result = self._preview_digital_pdf(pdf_path, defaults, max_width)
        if result is None:
            result = self._preview_ocr_pdf(pdf_path, defaults, max_width)
        return result

    def _preview_digital_single(self, pdf_path, defaults, page_num, max_width):
        """Render a single page using PyMuPDF."""
        if fitz is None:
            return None
        try:
            doc = fitz.open(pdf_path)
            total = len(doc)
            if page_num >= total:
                doc.close()
                return None

            page = doc[page_num]
            # Fall back to OCR if the page has no extractable text (scanned)
            if len(page.get_text().strip()) < 50:
                doc.close()
                return None

            page_rect = page.rect
            zoom = max_width / page_rect.width
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)

            pil_img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            buf = io.BytesIO()
            pil_img.save(buf, format="JPEG", quality=85)
            img_b64 = base64.b64encode(buf.getvalue()).decode("ascii")

            highlights = self._find_highlights_digital(page, defaults, zoom)
            capped_total = min(total, self.MAX_PREVIEW_PAGES)
            doc.close()

            return {
                "page": {
                    "image": f"data:image/jpeg;base64,{img_b64}",
                    "width": pix.width,
                    "height": pix.height,
                    "highlights": highlights,
                    "page": page_num + 1,
                },
                "total_pages": capped_total,
            }
        except Exception:
            return None

    def _preview_ocr_single(self, pdf_path, defaults, page_num, max_width):
        """Render a single page using pdf2image + pytesseract."""
        if convert_from_path is None or pytesseract is None or Image is None:
            return None
        try:
            images = convert_from_path(
                pdf_path,
                first_page=page_num + 1,
                last_page=page_num + 1,
                dpi=200,
                poppler_path=_POPPLER_PATH,
            )
            if not images:
                return None

            # Get total page count
            if fitz is not None:
                doc = fitz.open(pdf_path)
                total = min(len(doc), self.MAX_PREVIEW_PAGES)
                doc.close()
            else:
                total = page_num + 1  # can't know total without fitz

            orig_img = images[0]
            img = orig_img.copy()
            orig_w, orig_h = img.size
            if orig_w > max_width:
                scale = max_width / orig_w
                new_h = int(orig_h * scale)
                img = img.resize((max_width, new_h), Image.LANCZOS)
            img_width, img_height = img.size
            scale_x = img_width / orig_w
            scale_y = img_height / orig_h

            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            img_b64 = base64.b64encode(buf.getvalue()).decode("ascii")

            # OCR language detection
            ocr_lang = "spa+eng"
            try:
                available = pytesseract.get_languages()
                if "spa" not in available:
                    ocr_lang = "eng"
            except Exception:
                pass

            ocr_data = pytesseract.image_to_data(
                orig_img,
                lang=ocr_lang,
                output_type=pytesseract.Output.DICT,
                config="--psm 6",
            )
            highlights = self._find_highlights_ocr(ocr_data, defaults, scale_x, scale_y)

            return {
                "page": {
                    "image": f"data:image/jpeg;base64,{img_b64}",
                    "width": img_width,
                    "height": img_height,
                    "highlights": highlights,
                    "page": page_num + 1,
                },
                "total_pages": total,
            }
        except Exception:
            return None

    def _find_highlights_digital(self, page, defaults, zoom):
        """Find bounding boxes for field values in a digital PDF page."""
        highlights = []
        for field, value in defaults.items():
            if not value or field not in self._FIELD_COLORS:
                continue
            val_str = str(value)
            rects = page.search_for(val_str)
            if not rects:
                normalized = self._normalize_for_search(val_str)
                if normalized != val_str:
                    rects = page.search_for(normalized)
            # Fallback: search for first word only
            if not rects and " " in val_str:
                first_word = val_str.split()[0]
                rects = page.search_for(first_word)
            for rect in rects:
                w = (rect.x1 - rect.x0) * zoom
                h = (rect.y1 - rect.y0) * zoom
                if w > 1 and h > 1:  # Skip degenerate rectangles
                    highlights.append(
                        {
                            "field": field,
                            "color": self._FIELD_COLORS[field],
                            "x": rect.x0 * zoom,
                            "y": rect.y0 * zoom,
                            "w": w,
                            "h": h,
                        }
                    )
        return highlights

    def _find_highlights_ocr(self, ocr_data, defaults, scale_x, scale_y):
        """Find tight bounding boxes for field values in OCR data.

        Uses exact word-sequence matching so highlights cover only the value
        text, not the surrounding label or whitespace.
        """
        highlights = []
        words = ocr_data.get("text", [])
        lefts = ocr_data.get("left", [])
        tops = ocr_data.get("top", [])
        widths = ocr_data.get("width", [])
        heights = ocr_data.get("height", [])

        # Compact index: only non-empty word positions
        word_idx = [i for i in range(len(words)) if (words[i] or "").strip()]
        nw = len(word_idx)

        def _box(seq):
            x0 = min(lefts[q] for q in seq)
            y0 = min(tops[q] for q in seq)
            x1 = max(lefts[q] + widths[q] for q in seq)
            y1 = max(tops[q] + heights[q] for q in seq)
            return (x0, y0, x1, y1) if (x1 - x0) > 2 and (y1 - y0) > 2 else None

        for field, value in defaults.items():
            if not value or field not in self._FIELD_COLORS:
                continue
            value_str = str(value).strip().upper()
            value_wds = value_str.split()
            if not value_wds:
                continue
            vw = len(value_wds)
            value_norms = [self._normalize_for_search(w) for w in value_wds]

            found_box = None

            # Strategy 1: exact consecutive word-sequence match (tight box)
            for start in range(nw - vw + 1):
                seq = word_idx[start : start + vw]
                if all(
                    self._normalize_for_search((words[seq[k]] or "").strip().upper())
                    == value_norms[k]
                    for k in range(vw)
                ):
                    found_box = _box(seq)
                    break

            # Strategy 2: value as substring of a word window — trim to value words
            if not found_box:
                val_norm = self._normalize_for_search(value_str)
                for start in range(nw):
                    concat = ""
                    seq = []
                    for k in range(start, min(start + vw + 8, nw)):
                        wk = (words[word_idx[k]] or "").strip()
                        if not wk:
                            continue
                        concat = (concat + " " + wk).strip() if concat else wk
                        seq.append(word_idx[k])
                        cu = concat.upper()
                        cn = self._normalize_for_search(cu)
                        if value_str in cu or val_norm in cn:
                            # Find where in concat the value starts
                            pos = cu.find(value_str)
                            if pos < 0:
                                pos = cn.find(val_norm)
                            # Skip prefix words up to pos
                            trim_start, chars = 0, 0
                            for t, wi in enumerate(seq):
                                if chars >= pos:
                                    trim_start = t
                                    break
                                chars += len((words[wi] or "").strip()) + 1
                            # Include only enough words to cover the value
                            trim_end, chars = len(seq), 0
                            for t, wi in enumerate(seq[trim_start:]):
                                chars += len((words[wi] or "").strip()) + 1
                                if chars >= len(value_str):
                                    trim_end = trim_start + t + 1
                                    break
                            trimmed = seq[trim_start:trim_end]
                            if trimmed:
                                found_box = _box(trimmed)
                            break

            # Strategy 3: first word only (last resort for multi-word values)
            if not found_box and vw >= 2:
                fw_norm = value_norms[0]
                for wi in word_idx:
                    if (
                        self._normalize_for_search((words[wi] or "").strip().upper())
                        == fw_norm
                    ):
                        found_box = _box([wi])
                        break

            if found_box:
                x0, y0, x1, y1 = found_box
                highlights.append(
                    {
                        "field": field,
                        "color": self._FIELD_COLORS[field],
                        "x": x0 * scale_x,
                        "y": y0 * scale_y,
                        "w": (x1 - x0) * scale_x,
                        "h": (y1 - y0) * scale_y,
                    }
                )

        return highlights

    @staticmethod
    def _normalize_for_search(value):
        """Normalize a string for fuzzy matching: strip accents for comparison."""
        import unicodedata

        nfkd = unicodedata.normalize("NFKD", str(value))
        return "".join(c for c in nfkd if not unicodedata.combining(c))

    def _preview_digital_pdf(self, pdf_path, defaults, max_width):
        """Use PyMuPDF to render all pages and search for field values."""
        if fitz is None:
            return None
        try:
            doc = fitz.open(pdf_path)
            total = len(doc)
            if total == 0:
                doc.close()
                return None

            # Check first page for enough text (digital PDF check)
            first_text = doc[0].get_text()
            if len(first_text.strip()) < 50:
                doc.close()
                return None

            pages = []
            render_count = min(total, self.MAX_PREVIEW_PAGES)
            for page_num in range(render_count):
                page = doc[page_num]
                page_rect = page.rect
                zoom = max_width / page_rect.width
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                # Convert to JPEG for smaller payload
                pil_img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                buf = io.BytesIO()
                pil_img.save(buf, format="JPEG", quality=75)
                img_bytes = buf.getvalue()
                img_b64 = base64.b64encode(img_bytes).decode("ascii")

                # Find bounding boxes for each field value
                highlights = []
                for field, value in defaults.items():
                    if not value or field not in self._FIELD_COLORS:
                        continue
                    val_str = str(value)
                    # Try exact search first, then without accents (ñ -> n fallback)
                    rects = page.search_for(val_str)
                    if not rects:
                        normalized = self._normalize_for_search(val_str)
                        if normalized != val_str:
                            rects = page.search_for(normalized)
                    for rect in rects:
                        highlights.append(
                            {
                                "field": field,
                                "color": self._FIELD_COLORS[field],
                                "x": rect.x0 * zoom,
                                "y": rect.y0 * zoom,
                                "w": (rect.x1 - rect.x0) * zoom,
                                "h": (rect.y1 - rect.y0) * zoom,
                            }
                        )

                pages.append(
                    {
                        "image": f"data:image/jpeg;base64,{img_b64}",
                        "width": pix.width,
                        "height": pix.height,
                        "highlights": highlights,
                        "page": page_num + 1,
                    }
                )

            doc.close()
            return {"pages": pages, "total_pages": total}
        except Exception:
            return None

    def _preview_ocr_pdf(self, pdf_path, defaults, max_width):
        """Use pdf2image + pytesseract for scanned PDFs."""
        if convert_from_path is None or pytesseract is None or Image is None:
            return None
        try:
            images = convert_from_path(
                pdf_path,
                first_page=1,
                last_page=self.MAX_PREVIEW_PAGES,
                poppler_path=_POPPLER_PATH,
            )
            if not images:
                return None

            # Detect OCR language
            ocr_lang = "spa+eng"
            try:
                available = pytesseract.get_languages()
                if "spa" not in available:
                    ocr_lang = "eng"
            except Exception:
                pass

            pages = []
            for page_num, orig_img in enumerate(images):
                img = orig_img.copy()
                orig_w, orig_h = img.size
                if orig_w > max_width:
                    scale = max_width / orig_w
                    new_h = int(orig_h * scale)
                    img = img.resize((max_width, new_h), Image.LANCZOS)
                img_width, img_height = img.size
                scale_x = img_width / orig_w
                scale_y = img_height / orig_h

                buf = io.BytesIO()
                img.save(buf, format="JPEG", quality=75)
                img_bytes = buf.getvalue()
                img_b64 = base64.b64encode(img_bytes).decode("ascii")

                # OCR with bounding box data on original-size image
                ocr_data = pytesseract.image_to_data(
                    orig_img,
                    lang=ocr_lang,
                    output_type=pytesseract.Output.DICT,
                )

                highlights = []
                for field, value in defaults.items():
                    if not value or field not in self._FIELD_COLORS:
                        continue
                    value_str = str(value).upper()
                    value_norm = self._normalize_for_search(value_str)
                    words = ocr_data.get("text", [])
                    n = len(words)
                    found = False
                    for i in range(n):
                        if found:
                            break
                        if not words[i]:
                            continue
                        concat = ""
                        for j in range(i, min(i + 10, n)):
                            w = (words[j] or "").strip()
                            if not w:
                                continue
                            concat = (concat + " " + w).strip() if concat else w
                            concat_upper = concat.upper()
                            concat_norm = self._normalize_for_search(concat_upper)
                            if value_str in concat_upper or value_norm in concat_norm:
                                x0 = ocr_data["left"][i]
                                y0 = ocr_data["top"][i]
                                x1 = ocr_data["left"][j] + ocr_data["width"][j]
                                y1 = max(
                                    ocr_data["top"][k] + ocr_data["height"][k]
                                    for k in range(i, j + 1)
                                    if (words[k] or "").strip()
                                )
                                highlights.append(
                                    {
                                        "field": field,
                                        "color": self._FIELD_COLORS[field],
                                        "x": x0 * scale_x,
                                        "y": y0 * scale_y,
                                        "w": (x1 - x0) * scale_x,
                                        "h": (y1 - y0) * scale_y,
                                    }
                                )
                                found = True
                                break

                pages.append(
                    {
                        "image": f"data:image/jpeg;base64,{img_b64}",
                        "width": img_width,
                        "height": img_height,
                        "highlights": highlights,
                        "page": page_num + 1,
                    }
                )

            return {"pages": pages, "total_pages": len(images)}
        except Exception:
            return None

    def generate_filename_parts(self, pdf_path, verbose=False, original_filename=None):
        """Extrae información y genera partes del nombre de archivo"""
        # Extrae texto del PDF
        text = self.extract_text_from_pdf(pdf_path, verbose=verbose)

        # Si no hay texto suficiente, intenta extraer del nombre del archivo
        filename_data = {}
        if original_filename:
            filename_data = self.extract_from_filename(original_filename)
            if verbose:
                print(f"  📁 Información extraída del nombre: {filename_data}")

        # Si no hay texto del PDF, usa solo el nombre del archivo
        if not text or len(text.strip()) < 10:
            if filename_data:
                # Construye nombre desde el nombre del archivo
                parts = []
                metadata = {}

                # Orden específico: DNI, NOMBRE, EMPRESA, TIPO_EXAMEN,
                # CMESPINAR, FECHA

                # 1. DNI (requerido)
                if filename_data.get("dni"):
                    parts.append(filename_data["dni"])
                    metadata["dni"] = filename_data["dni"]
                else:
                    # Sin DNI no podemos generar nombre
                    if verbose:
                        print("  [WARN] DNI no encontrado en nombre del archivo")
                    return None, None

                # 2. NOMBRE
                if filename_data.get("person_name"):
                    name_clean = "_".join(filename_data["person_name"].split()[:4])
                    name_clean = re.sub(r'[<>:"/\\|?*&]', "", name_clean)
                    parts.append(name_clean)
                    metadata["nombre"] = filename_data["person_name"]

                # 3. EMPRESA
                if filename_data.get("company"):
                    company_clean = "_".join(filename_data["company"].split()[:4])
                    company_clean = re.sub(r'[<>:"/\\|?*&]', "", company_clean)
                    parts.append(company_clean)
                    metadata["empresa"] = filename_data["company"]

                # 4. TIPO DE EXAMEN
                if filename_data.get("exam_type"):
                    parts.append(filename_data["exam_type"])
                    metadata["tipo_examen"] = filename_data["exam_type"]

                # 5. CMESPINAR (constante)
                parts.append("CMESPINAR")
                metadata["centro"] = "CMESPINAR"

                # 6. FECHA DE EVALUACION
                if filename_data.get("date"):
                    date_clean = (
                        filename_data["date"].replace(".", "-").replace("/", "-")
                    )
                    parts.append(date_clean)
                    metadata["fecha"] = filename_data["date"]

                if parts:
                    new_name = "_".join(parts)
                    new_name = re.sub(r'[<>:"/\\|?*]', "", new_name)
                    new_name = re.sub(r"\s+", "_", new_name)
                    new_name = re.sub(r"_+", "_", new_name)
                    new_name = new_name.strip("_")
                    return new_name + ".pdf", metadata

            # Si tampoco hay datos del nombre, retorna None
            if verbose:
                print(f"  [WARN] Texto extraído: {len(text) if text else 0} caracteres")
            return None, None

        # Extrae información del texto del PDF
        date = self.extract_dates(text)
        number = self.extract_numbers(text)
        entity = self.extract_entity_names(text)
        doc_type = self.detect_document_type(text)
        exam_type = self.extract_exam_type(text)

        # Si no se encontró información en el texto, usa datos del nombre del archivo
        # como fallback
        if not date and filename_data.get("date"):
            date = filename_data["date"]
        if not number and filename_data.get("dni"):
            number = filename_data["dni"]
        if not doc_type and filename_data.get("doc_type"):
            doc_type = filename_data["doc_type"]
        if not exam_type and filename_data.get("exam_type"):
            exam_type = filename_data["exam_type"]
        if not entity:
            if filename_data.get("person_name"):
                entity = filename_data["person_name"]
            elif filename_data.get("company"):
                entity = filename_data["company"]

        # Construye las partes en el orden específico requerido:
        # 1. DNI
        # 2. NOMBRE
        # 3. EMPRESA
        # 4. TIPO DE EXAMEN
        # 5. CMESPINAR (constante)
        # 6. FECHA DE EVALUACION
        parts = []
        metadata = {}

        # 1. DNI
        dni_value = None
        if number:
            # Si es DNI (8 dígitos) o contiene DNI en el texto
            if "DNI" in text.upper() or (number and len(number) == 8):
                dni_value = number
            elif number:
                dni_value = number

        # También verifica en filename_data
        if not dni_value and filename_data.get("dni"):
            dni_value = filename_data["dni"]

        if dni_value:
            parts.append(dni_value)
            metadata["dni"] = dni_value
            if verbose:
                print(f"  🔢 DNI: {dni_value}")
        else:
            # Si no hay DNI, no podemos generar el nombre
            if verbose:
                print("  [WARN] DNI no encontrado - requerido para generar nombre")
            return None, None

        # 2. NOMBRE (persona)
        person_name = None
        if entity:
            # Verifica si es un nombre de persona (típicamente 3-4 palabras en mayúsculas)
            entity_words = entity.split()
            if len(entity_words) >= 2:
                # Si tiene 2-4 palabras y son mayúsculas, probablemente es nombre de persona
                if all(
                    word[0].isupper() if word else False for word in entity_words[:3]
                ):
                    person_name = entity

        # También verifica en filename_data
        if not person_name and filename_data.get("person_name"):
            person_name = filename_data["person_name"]

        if person_name:
            # Limpia y normaliza el nombre
            name_clean = "_".join(person_name.split()[:4])  # Máximo 4 palabras
            name_clean = re.sub(r'[<>:"/\\|?*&]', "", name_clean)
            parts.append(name_clean)
            metadata["nombre"] = person_name
            if verbose:
                print(f"  👤 Nombre: {person_name}")
        else:
            # Nombre es opcional pero recomendado
            if verbose:
                print("  [WARN] Nombre no encontrado")

        # 3. EMPRESA
        company_name = None
        if filename_data.get("company"):
            company_name = filename_data["company"]
        elif entity and not person_name:
            # Si entity no es nombre de persona, podría ser empresa
            company_name = entity

        if company_name:
            # Limpia y normaliza la empresa
            company_clean = "_".join(company_name.split()[:4])  # Máximo 4 palabras
            company_clean = re.sub(r'[<>:"/\\|?*&]', "", company_clean)
            parts.append(company_clean)
            metadata["empresa"] = company_name
            if verbose:
                print(f"  🏢 Empresa: {company_name}")
        else:
            # Empresa es opcional
            if verbose:
                print("  [WARN] Empresa no encontrada")

        # 4. TIPO DE EXAMEN
        exam_type_value = None
        if exam_type:
            exam_type_value = exam_type
        elif filename_data.get("exam_type"):
            exam_type_value = filename_data["exam_type"]

        if exam_type_value:
            parts.append(exam_type_value)
            metadata["tipo_examen"] = exam_type_value
            if verbose:
                print(f"  🏥 Tipo de examen: {exam_type_value}")
        else:
            # Tipo de examen es opcional
            if verbose:
                print("  [WARN] Tipo de examen no encontrado")

        # 5. CMESPINAR (constante)
        parts.append("CMESPINAR")
        metadata["centro"] = "CMESPINAR"
        if verbose:
            print(f"  🏥 Centro: CMESPINAR")

        # 6. FECHA DE EVALUACION
        if date:
            try:
                # Normaliza fecha: convierte 31.12.25 a 31-12-25, o 31/12/2025 a 31-12-2025
                date_clean = re.sub(r"[^\d./-]", "", date)
                # Si tiene formato DD.MM.YY, convierte a DD-MM-YY
                if "." in date_clean:
                    date_clean = date_clean.replace(".", "-")
                else:
                    date_clean = date_clean.replace("/", "-")
                parts.append(date_clean)
                metadata["fecha"] = date
                if verbose:
                    print(f"  📅 Fecha: {date}")
            except Exception:
                pass
        elif filename_data.get("date"):
            date_clean = filename_data["date"].replace(".", "-").replace("/", "-")
            parts.append(date_clean)
            metadata["fecha"] = filename_data["date"]
            if verbose:
                print(f"  📅 Fecha (del nombre): {filename_data['date']}")

        # Valida que al menos tengamos DNI y FECHA (mínimos requeridos)
        if not parts or len(parts) < 2:
            if verbose:
                print("  [WARN] Información insuficiente para generar nombre")
            return None, None

        # Une las partes y limpia caracteres no válidos
        new_name = "_".join(parts)
        new_name = re.sub(r'[<>:"/\\|?*]', "", new_name)
        new_name = re.sub(r"\s+", "_", new_name)
        new_name = re.sub(r"_+", "_", new_name)
        new_name = new_name.strip("_")  # Elimina guiones al inicio/final

        return new_name + ".pdf", metadata
